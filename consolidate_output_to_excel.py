#!/usr/bin/env python3
"""
data/output 内の各フォルダにあるTSVを集約し、1つのExcelにまとめる。

- フォルダごとにシートを作成（シート名 = フォルダ名）
- 同一行（LineNo in file が同じ）は1行にまとめる
- 重複時は「一番先に出現したファイル」の検索ワード（code）を優先
- レポート.md は対象外。*.tsv のみ処理

使い方:
  python consolidate_output_to_excel.py              # 全フォルダを集約
  python consolidate_output_to_excel.py --only T1572 # T1572 で始まるフォルダだけ集約
"""

import csv
import re
import sys
from pathlib import Path

# Excel シート名: 31文字まで、\ / : * ? [ ] は使えない
def sanitize_sheet_name(name: str, max_len: int = 31) -> str:
    s = re.sub(r'[\\/:*?\[\]]', "_", name)
    if len(s) > max_len:
        s = s[:max_len]
    return s or "Sheet"


def load_tsv(path: Path) -> list[dict]:
    """1つのTSVを読み、行のリストで返す。"""
    rows = []
    for enc in ("utf-8-sig", "utf-8", "cp932", "shift_jis"):
        try:
            with open(path, encoding=enc, newline="") as f:
                r = csv.DictReader(f, delimiter="\t")
                for row in r:
                    rows.append(dict(row))
            return rows
        except (UnicodeDecodeError, csv.Error):
            rows = []
    return rows


def merge_tsvs_in_folder(folder: Path) -> tuple[list[dict], list[str]]:
    """
    フォルダ内の全TSVを読み、同一行（LineNo in file）は1行にまとめる。
    先に出現したファイルの code を優先するため、TSVはファイル名でソートして読む。
    戻り値: (行リスト, カラム名リスト)
    """
    tsv_files = sorted(folder.iterdir(), key=lambda p: p.name)
    tsv_files = [p for p in tsv_files if p.suffix.lower() == ".tsv"]
    if not tsv_files:
        return [], []

    seen_line_no: set[str] = set()
    merged_rows: list[dict] = []
    columns: list[str] = []

    for tsv_path in tsv_files:
        rows = load_tsv(tsv_path)
        if not rows:
            continue
        if not columns:
            columns = list(rows[0].keys())
        for row in rows:
            line_no = row.get("LineNo in file", "").strip()
            key = line_no
            if key in seen_line_no:
                continue
            seen_line_no.add(key)
            merged_rows.append(row)

    return merged_rows, columns


def main() -> int:
    import argparse
    ap = argparse.ArgumentParser(description="data/output 内のTSVを1つのExcelに集約")
    ap.add_argument(
        "--only", "-o",
        metavar="PREFIX",
        default=None,
        help="このプレフィックスで始まるフォルダだけ集約（例: --only T1572）",
    )
    args = ap.parse_args()

    base_dir = Path(__file__).resolve().parent
    output_root = base_dir / "data" / "output"

    if not output_root.exists() or not output_root.is_dir():
        print(f"エラー: {output_root} が存在しません。", file=sys.stderr)
        return 1

    # サブフォルダのみ（TSVが入っているフォルダ）。--only のときはプレフィックスで絞る
    folders = [
        p for p in output_root.iterdir()
        if p.is_dir() and not p.name.startswith(".")
    ]
    if args.only:
        prefix = args.only.strip()
        folders = [p for p in folders if p.name.startswith(prefix)]
    folders = sorted(folders)
    if not folders:
        if args.only:
            print(f"エラー: {output_root} に「{args.only}」で始まるフォルダがありません。", file=sys.stderr)
        else:
            print(f"エラー: {output_root} にサブフォルダがありません。", file=sys.stderr)
        return 1

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment
    except ImportError:
        print("エラー: openpyxl がありません。pip install openpyxl を実行してください。", file=sys.stderr)
        return 1

    def fit_column_widths(ws, num_cols, min_w=8, max_w=80):
        """各列の内容に合わせて幅を設定。長い列は max_w で打ち止め。"""
        for c in range(1, num_cols + 1):
            max_len = 0
            col_letter = get_column_letter(c)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=c, max_col=c):
                for cell in row:
                    if cell.value is None:
                        continue
                    s = str(cell.value)
                    # 改行は最長行で評価
                    for line in s.split("\n"):
                        # 日本語等は2文字で1幅とみなす簡易換算
                        w = 0
                        for ch in line:
                            w += 2 if ord(ch) > 127 else 1
                        max_len = max(max_len, min(w, max_w * 2))
            width = min(max(min_w, max_len // 2 + 1), max_w)
            ws.column_dimensions[col_letter].width = width

    wb = Workbook()
    first = True
    for folder in folders:
        merged, columns = merge_tsvs_in_folder(folder)
        if not columns:
            print(f"  スキップ（TSVなし）: {folder.name}", file=sys.stderr)
            continue
        if len(merged) == 0:
            print(f"  スキップ（ログ0件）: {folder.name}", file=sys.stderr)
            continue

        sheet_name = sanitize_sheet_name(folder.name)
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for c, col in enumerate(columns, start=1):
            ws.cell(row=1, column=c, value=col)
        for r, row in enumerate(merged, start=2):
            for c, col in enumerate(columns, start=1):
                val = row.get(col, "")
                if val is not None:
                    cell = ws.cell(row=r, column=c, value=val)
                    # 長いセルは折り返して表示
                    if isinstance(val, str) and len(val) > 60:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
        fit_column_widths(ws, len(columns))
        ws.freeze_panes = "A2"  # 1行目を固定
        print(f"  {folder.name}: {len(merged)} 行（重複除く）", file=sys.stderr)

    out_path = output_root / "consolidated.xlsx"
    wb.save(out_path)
    print(f"出力: {out_path}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
